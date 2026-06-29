from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Count
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.db.models import Q
import json
from .models import Division, EmployeeEmail
from .forms import AccountCreateForm, AccountUpdateForm, EmployeeEmailForm
from apps.inventory.models import Equipment
from apps.tickets.models import Ticket

def is_admin(user):
    return user.is_staff or user.is_superuser

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

class DashboardView(LoginRequiredMixin, ListView):
    template_name = 'dashboard.html'
    context_object_name = 'recent_tickets'
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return Ticket.objects.all()[:5]
        return Ticket.objects.filter(reporter=self.request.user)[:5]
    
    def get_context_data(self, **kwargs):
        from django.utils import timezone
        from django.db.models import F, Sum
        from apps.inventory.models import EquipmentCategory
        from apps.atk.models import ATKItem, ATKRequest
        context = super().get_context_data(**kwargs)
        context['today'] = timezone.now()
        if self.request.user.is_staff:
            # Total counts
            context['total_users'] = EmployeeEmail.objects.count()
            context['pending_tickets'] = Ticket.objects.filter(status='new').count()
            context['pending_atk_requests'] = ATKRequest.objects.filter(status='pending').count()
            
            email_counts = EmployeeEmail.objects.aggregate(
                active=Count('id', filter=Q(is_active=True)),
                inactive=Count('id', filter=Q(is_active=False)),
            )
            context['email_active'] = email_counts['active']
            context['email_inactive'] = email_counts['inactive']
            
            # Get main categories
            elektronik_category = EquipmentCategory.objects.filter(name='Elektronik').first()
            furniture_category = EquipmentCategory.objects.filter(name='Furniture').first()
            
            # Elektronik breakdown
            if elektronik_category:
                elektronik_subcategories = elektronik_category.get_children()
                elektronik_equipment = Equipment.objects.filter(category__in=elektronik_subcategories)
                elektronik_counts = elektronik_equipment.aggregate(
                    total=Count('id'),
                    available=Count('id', filter=Q(status='available')),
                    borrowed=Count('id', filter=Q(status='borrowed')),
                    broken=Count('id', filter=Q(status='broken')),
                    service=Count('id', filter=Q(status='service')),
                )
                context['total_elektronik'] = elektronik_counts['total']
                context['elektronik_available'] = elektronik_counts['available']
                context['elektronik_borrowed'] = elektronik_counts['borrowed']
                context['elektronik_broken'] = elektronik_counts['broken']
                context['elektronik_service'] = elektronik_counts['service']
            else:
                context['total_elektronik'] = 0
                context['elektronik_available'] = 0
                context['elektronik_borrowed'] = 0
                context['elektronik_broken'] = 0
                context['elektronik_service'] = 0
            
            # Furniture breakdown
            if furniture_category:
                furniture_subcategories = furniture_category.get_children()
                furniture_equipment = Equipment.objects.filter(category__in=furniture_subcategories)
                furniture_counts = furniture_equipment.aggregate(
                    total=Count('id'),
                    available=Count('id', filter=Q(status='available')),
                    borrowed=Count('id', filter=Q(status='borrowed')),
                    broken=Count('id', filter=Q(status='broken')),
                    service=Count('id', filter=Q(status='service')),
                )
                context['total_furniture'] = furniture_counts['total']
                context['furniture_available'] = furniture_counts['available']
                context['furniture_borrowed'] = furniture_counts['borrowed']
                context['furniture_broken'] = furniture_counts['broken']
                context['furniture_service'] = furniture_counts['service']
            else:
                context['total_furniture'] = 0
                context['furniture_available'] = 0
                context['furniture_borrowed'] = 0
                context['furniture_broken'] = 0
                context['furniture_service'] = 0
            
            ticket_counts = Ticket.objects.aggregate(
                new=Count('id', filter=Q(status='new')),
                in_progress=Count('id', filter=Q(status='in_progress')),
                completed=Count('id', filter=Q(status='completed')),
                cancelled=Count('id', filter=Q(status='cancelled')),
            )
            context['ticket_new'] = ticket_counts['new']
            context['ticket_in_progress'] = ticket_counts['in_progress']
            context['ticket_completed'] = ticket_counts['completed']
            context['ticket_cancelled'] = ticket_counts['cancelled']

            # ATK stock overview for dashboard chart
            atk_items = ATKItem.objects.select_related('category')
            context['atk_total_items'] = atk_items.count()
            context['atk_total_stock'] = atk_items.aggregate(total=Sum('current_stock'))['total'] or 0
            context['atk_low_stock'] = atk_items.filter(
                current_stock__gt=0,
                current_stock__lte=F('minimum_stock')
            ).count()
            context['atk_out_of_stock'] = atk_items.filter(current_stock=0).count()

            atk_category_rows = (
                atk_items.values('category__name')
                .annotate(total_stock=Sum('current_stock'), item_count=Count('id'))
                .order_by('-total_stock', 'category__name')[:6]
            )
            max_stock = max([row['total_stock'] or 0 for row in atk_category_rows], default=0)
            context['atk_category_chart'] = [
                {
                    'name': row['category__name'] or 'Tanpa Kategori',
                    'stock': row['total_stock'] or 0,
                    'items': row['item_count'],
                    'percent': round(((row['total_stock'] or 0) / max_stock) * 100) if max_stock else 0,
                }
                for row in atk_category_rows
            ]
        else:
            my_tickets = Ticket.objects.filter(reporter=self.request.user)
            my_atk_requests = ATKRequest.objects.filter(
                requester=self.request.user
            ).select_related('item').prefetch_related('lines__item')
            context['my_tickets'] = my_tickets.count()
            context['my_open_tickets'] = my_tickets.filter(
                status__in=['new', 'in_progress']
            ).count()
            context['my_atk_requests'] = my_atk_requests.count()
            context['my_atk_total_quantity'] = sum(atk_request.total_quantity for atk_request in my_atk_requests)
            context['my_pending_atk_requests'] = my_atk_requests.filter(
                status='pending'
            ).count()
            context['my_approved_atk_requests'] = my_atk_requests.filter(
                status='approved'
            ).count()
            context['my_recent_atk_requests'] = my_atk_requests[:5]
            context['my_recent_tickets'] = my_tickets[:5]
        return context


class AccountListView(AdminRequiredMixin, ListView):
    model = User
    template_name = 'users/account_list.html'
    context_object_name = 'accounts'
    paginate_by = 15

    def get_queryset(self):
        queryset = User.objects.select_related('profile__division').order_by('username')

        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
                | Q(profile__division__name__icontains=search)
            )

        role = self.request.GET.get('role')
        if role == 'admin':
            queryset = queryset.filter(is_staff=True)
        elif role == 'user':
            queryset = queryset.filter(is_staff=False)

        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['divisions'] = Division.objects.all()
        return context


class AccountCreateView(AdminRequiredMixin, CreateView):
    model = User
    form_class = AccountCreateForm
    template_name = 'users/account_form.html'
    success_url = reverse_lazy('users:account_list')

    def form_valid(self, form):
        messages.success(self.request, 'Akun user berhasil dibuat.')
        return super().form_valid(form)


class AccountUpdateView(AdminRequiredMixin, UpdateView):
    model = User
    form_class = AccountUpdateForm
    template_name = 'users/account_form.html'
    success_url = reverse_lazy('users:account_list')

    def form_valid(self, form):
        if self.object == self.request.user:
            if not form.cleaned_data.get('is_staff') or not form.cleaned_data.get('is_active'):
                messages.error(self.request, 'Anda tidak bisa menonaktifkan atau mencabut role admin akun sendiri.')
                return self.form_invalid(form)
        messages.success(self.request, 'Akun user berhasil diupdate.')
        return super().form_valid(form)


class DivisionListView(AdminRequiredMixin, ListView):
    model = Division
    template_name = 'users/division_list.html'
    context_object_name = 'divisions'
    paginate_by = 20

    def get_queryset(self):
        queryset = Division.objects.annotate(member_count=Count('members')).order_by('name')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(description__icontains=search))
        return queryset


class DivisionCreateView(AdminRequiredMixin, CreateView):
    model = Division
    template_name = 'users/division_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('users:division_list')

    def form_valid(self, form):
        messages.success(self.request, 'Divisi berhasil ditambahkan.')
        return super().form_valid(form)


class DivisionUpdateView(AdminRequiredMixin, UpdateView):
    model = Division
    template_name = 'users/division_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('users:division_list')

    def form_valid(self, form):
        messages.success(self.request, 'Divisi berhasil diupdate.')
        return super().form_valid(form)

# Email Management Views
class EmployeeEmailListView(AdminRequiredMixin, ListView):
    model = EmployeeEmail
    template_name = 'users/email_list.html'
    context_object_name = 'emails'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                full_name__icontains=search
            ) | queryset.filter(
                primary_email__icontains=search
            ) | queryset.filter(
                employee_id__icontains=search
            )
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(is_active=(status == 'active'))
        
        return queryset

class EmployeeEmailCreateView(AdminRequiredMixin, CreateView):
    model = EmployeeEmail
    form_class = EmployeeEmailForm
    template_name = 'users/email_form.html'
    success_url = reverse_lazy('users:email_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Email karyawan berhasil ditambahkan!')
        return super().form_valid(form)

class EmployeeEmailUpdateView(AdminRequiredMixin, UpdateView):
    model = EmployeeEmail
    form_class = EmployeeEmailForm
    template_name = 'users/email_form.html'
    success_url = reverse_lazy('users:email_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Email karyawan berhasil diupdate!')
        return super().form_valid(form)

class EmployeeEmailDeleteView(AdminRequiredMixin, DeleteView):
    model = EmployeeEmail
    template_name = 'users/email_confirm_delete.html'
    success_url = reverse_lazy('users:email_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Email karyawan berhasil dihapus!')
        return super().delete(request, *args, **kwargs)


@require_POST
@login_required
def verify_admin_password(request):
    """Verify admin password untuk melihat password email"""
    try:
        data = json.loads(request.body)
        password = data.get('password')
        
        # Authenticate user
        user = authenticate(username=request.user.username, password=password)
        
        if user is not None and user.is_staff:
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Password salah'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

@login_required
def export_emails_excel(request):
    """Export email karyawan ke Excel (tanpa password)"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses untuk export data!')
        return redirect('users:email_list')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Karyawan"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['No', 'Nama Lengkap', 'NIK', 'Email Utama', 'Recovery Email', 'Recovery Phone', 'Status', 'Tanggal Dibuat']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    emails = EmployeeEmail.objects.all().order_by('-created_at')
    for row, email in enumerate(emails, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=email.full_name)
        ws.cell(row=row, column=3, value=email.employee_id)
        ws.cell(row=row, column=4, value=email.primary_email)
        ws.cell(row=row, column=5, value=email.recovery_email or '-')
        ws.cell(row=row, column=6, value=email.recovery_phone or '-')
        ws.cell(row=row, column=7, value='Aktif' if email.is_active else 'Nonaktif')
        ws.cell(row=row, column=8, value=email.created_at.strftime('%d/%m/%Y %H:%M'))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Email_Karyawan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
