from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Ticket
from apps.users.views import AdminRequiredMixin

class TicketListView(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'tickets/ticket_list.html'
    context_object_name = 'tickets'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
            queryset = queryset.filter(reporter=self.request.user)
        
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(title__icontains=search)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset

class TicketCreateView(LoginRequiredMixin, CreateView):
    model = Ticket
    template_name = 'tickets/ticket_form.html'
    fields = ['title', 'equipment', 'description', 'priority', 'attachment']
    success_url = reverse_lazy('tickets:ticket_list')

    def form_valid(self, form):
        form.instance.reporter = self.request.user
        messages.success(self.request, 'Tiket berhasil dibuat!')
        return super().form_valid(form)

class TicketDetailView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = 'tickets/ticket_detail.html'
    context_object_name = 'ticket'

@login_required
def update_ticket_status(request, pk):
    """View untuk update status dan notes tiket"""
    ticket = get_object_or_404(Ticket, pk=pk)
    
    # Only admin can update status
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses untuk mengubah status tiket!')
        return redirect('tickets:ticket_detail', pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        
        ticket.status = new_status
        ticket.notes = notes
        
        # Set completed_at jika status selesai atau tidak selesai
        if new_status in ['completed', 'cancelled']:
            ticket.completed_at = timezone.now()
        
        ticket.save()
        messages.success(request, f'Status tiket berhasil diubah menjadi "{ticket.get_status_display()}"!')
        return redirect('tickets:ticket_detail', pk=pk)
    
    return redirect('tickets:ticket_detail', pk=pk)

class TicketDeleteView(AdminRequiredMixin, DetailView):
    model = Ticket
    template_name = 'tickets/ticket_confirm_delete.html'
    context_object_name = 'ticket'
    
    def post(self, request, *args, **kwargs):
        ticket = self.get_object()
        ticket.delete()
        messages.success(request, 'Tiket berhasil dihapus!')
        return redirect('tickets:ticket_list')


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

@login_required
def export_tickets_excel(request):
    """Export tiket ke Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses untuk export data!')
        return redirect('tickets:ticket_list')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Tiket"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['No', 'Judul', 'Pelapor', 'Deskripsi', 'Prioritas', 'Status', 'Catatan', 'Tanggal Dibuat', 'Tanggal Selesai']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    tickets = Ticket.objects.select_related('reporter').all().order_by('-created_at')
    for row, ticket in enumerate(tickets, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=ticket.title)
        ws.cell(row=row, column=3, value=ticket.reporter.username)
        ws.cell(row=row, column=4, value=ticket.description)
        ws.cell(row=row, column=5, value=ticket.get_priority_display())
        ws.cell(row=row, column=6, value=ticket.get_status_display())
        ws.cell(row=row, column=7, value=ticket.notes or '-')
        ws.cell(row=row, column=8, value=ticket.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=9, value=ticket.completed_at.strftime('%d/%m/%Y %H:%M') if ticket.completed_at else '-')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Daftar_Tiket_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
