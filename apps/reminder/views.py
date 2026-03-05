from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from .models import NetworkCheckLog

@login_required
def reminder_view(request):
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        NetworkCheckLog.objects.create(checked_by=request.user, notes=notes)
        messages.success(request, 'Pengecekan jaringan berhasil dicatat!')
        return redirect('reminder:reminder')
    
    # Get filter parameter
    filter_by = request.GET.get('filter', 'all')
    
    # Base queryset
    logs = NetworkCheckLog.objects.all()
    
    # Apply filters
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())
    
    if filter_by == 'today':
        logs = logs.filter(checked_at__gte=today_start)
    elif filter_by == 'week':
        logs = logs.filter(checked_at__gte=week_start)
    
    # Statistics
    stats = {
        'today': NetworkCheckLog.objects.filter(checked_at__gte=today_start).count(),
        'week': NetworkCheckLog.objects.filter(checked_at__gte=week_start).count(),
        'total': NetworkCheckLog.objects.count(),
    }
    
    # Last check info
    last_check = NetworkCheckLog.objects.first()
    if last_check:
        time_diff = now - last_check.checked_at
        if time_diff.total_seconds() < 3600:  # Less than 1 hour
            stats['last_check'] = f"{int(time_diff.total_seconds() / 60)} menit yang lalu"
        elif time_diff.total_seconds() < 86400:  # Less than 1 day
            stats['last_check'] = f"{int(time_diff.total_seconds() / 3600)} jam yang lalu"
        else:
            stats['last_check'] = f"{time_diff.days} hari yang lalu"
        stats['last_check_by'] = last_check.checked_by.username
    else:
        stats['last_check'] = "Belum pernah dicek"
        stats['last_check_by'] = "-"
    
    # Check if already checked today
    checked_today = NetworkCheckLog.objects.filter(
        checked_by=request.user,
        checked_at__gte=today_start
    ).exists()
    
    # Pagination
    paginator = Paginator(logs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'logs': page_obj,
        'stats': stats,
        'checked_today': checked_today,
        'filter_by': filter_by,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }
    
    return render(request, 'reminder/reminder.html', context)


@login_required
def delete_log(request, pk):
    """Delete network check log (admin only)"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses untuk menghapus log!')
        return redirect('reminder:reminder')
    
    log = get_object_or_404(NetworkCheckLog, pk=pk)
    log.delete()
    messages.success(request, 'Log berhasil dihapus!')
    return redirect('reminder:reminder')


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

@login_required
def export_logs_excel(request):
    """Export network check logs to Excel"""
    if not request.user.is_staff:
        messages.error(request, 'Anda tidak memiliki akses untuk export data!')
        return redirect('reminder:reminder')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Pengecekan"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['No', 'Dicek Oleh', 'Waktu Pengecekan', 'Catatan']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    logs = NetworkCheckLog.objects.select_related('checked_by').all()
    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=log.checked_by.username)
        ws.cell(row=row, column=3, value=log.checked_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=4, value=log.notes or '-')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Riwayat_Pengecekan_Jaringan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
